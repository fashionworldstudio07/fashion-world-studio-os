"""Export routes — Excel and PDF report generation."""

import io
from datetime import datetime, timedelta, timezone, date
from typing import Annotated, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.api.deps import get_current_user
from app.models.user import User
from app.models.transaction import Transaction, TransactionService

router = APIRouter(prefix="/export", tags=["Export"])


@router.get("/transactions/excel")
async def export_transactions_excel(
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    start_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
):
    """Export transactions to Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = select(Transaction).order_by(Transaction.service_date.desc())
    if start_date:
        query = query.where(Transaction.service_date >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        query = query.where(Transaction.service_date <= datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1))

    result = await db.execute(query)
    transactions = result.scalars().all()

    # Load services for each transaction
    tx_services = {}
    for tx in transactions:
        sr = await db.execute(
            select(TransactionService).where(TransactionService.transaction_id == tx.id)
        )
        tx_services[tx.id] = sr.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Styling
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Fashion World Studio — Transactions Report"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="D4AF37")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws["A2"].font = Font(name="Calibri", size=9, color="888888")

    # Headers
    headers = ["Date", "Customer", "Services", "Amount (₹)", "Payment", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data
    for row, tx in enumerate(transactions, 5):
        services_list = tx_services.get(tx.id, [])
        services_str = ", ".join(s.service_name for s in services_list)

        ws.cell(row=row, column=1, value=tx.service_date.strftime("%d %b %Y") if tx.service_date else "").border = thin_border
        ws.cell(row=row, column=2, value=tx.customer.name if tx.customer else "Walk-in").border = thin_border
        ws.cell(row=row, column=3, value=services_str).border = thin_border
        ws.cell(row=row, column=4, value=float(tx.total_amount)).border = thin_border
        ws.cell(row=row, column=4).number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=(tx.payment_mode or "").upper()).border = thin_border
        ws.cell(row=row, column=6, value=tx.notes or "").border = thin_border

    # Summary row
    total_row = len(transactions) + 5
    ws.cell(row=total_row, column=3, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(float(tx.total_amount) for tx in transactions)).font = Font(bold=True)
    ws.cell(row=total_row, column=4).number_format = "#,##0.00"

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"FW_Transactions_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/transactions/pdf")
async def export_transactions_pdf(
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
):
    """Export transactions to PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    query = select(Transaction).order_by(Transaction.service_date.desc())
    if start_date:
        query = query.where(Transaction.service_date >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        query = query.where(Transaction.service_date <= datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1))

    result = await db.execute(query)
    transactions = result.scalars().all()

    tx_services = {}
    for tx in transactions:
        sr = await db.execute(
            select(TransactionService).where(TransactionService.transaction_id == tx.id)
        )
        tx_services[tx.id] = sr.scalars().all()

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#D4AF37"))
    elements.append(Paragraph("Fashion World Studio", title_style))
    elements.append(Paragraph(f"Transaction Report — {date.today().strftime('%d %b %Y')}", styles["Normal"]))
    elements.append(Spacer(1, 10*mm))

    # Table data
    data = [["Date", "Customer", "Services", "Amount (₹)", "Payment"]]
    total = 0.0
    for tx in transactions:
        svcs = tx_services.get(tx.id, [])
        svcs_str = ", ".join(s.service_name for s in svcs)
        amt = float(tx.total_amount)
        total += amt
        data.append([
            tx.service_date.strftime("%d/%m/%y") if tx.service_date else "—",
            tx.customer.name if tx.customer else "Walk-in",
            svcs_str[:40],
            f"₹{amt:,.0f}",
            (tx.payment_mode or "").upper(),
        ])
    data.append(["", "", "TOTAL", f"₹{total:,.0f}", ""])

    table = Table(data, colWidths=[60, 100, 150, 70, 60])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D4AF37")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.HexColor("#F8F8F8"), colors.white]),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)

    filename = f"FW_Report_{date.today().isoformat()}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
